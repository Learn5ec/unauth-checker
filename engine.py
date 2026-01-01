import json
import csv
import sys
import os
import re
import requests
from urllib.parse import urlparse
from typing import Dict, Any, List, Tuple, Optional
from ai_agent import AIAgent

# Try to import openpyxl for Excel generation, make it optional
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Lazy initialization to avoid import-time failures
_agent: Optional[AIAgent] = None

def get_agent() -> AIAgent:
    global _agent
    if _agent is None:
        _agent = AIAgent()
    return _agent

CSV_FIELDS = [
    "endpoint",
    "method",
    "params_count",
    "params_values",
    "status_codes",
    "response",
    "confidence",
    "confidence_level",
    "notes"
]

# -------------------------------------------------
# Progress Bar
# -------------------------------------------------

def update_progress(current: int, total: int):
    bar_len = 25
    filled = int(bar_len * current / total)
    bar = "█" * filled + "-" * (bar_len - filled)
    sys.stdout.write(f"\rProgress: [{bar}] {current} / {total} endpoints evaluated")
    sys.stdout.flush()

# -------------------------------------------------
# Excel Helpers (Formatted output)
# -------------------------------------------------

# Column width mappings (in cm, converted to Excel units)
# Excel column width: 1 cm ≈ 3.78 characters
COLUMN_WIDTHS_CM = {
    "endpoint": 4.7,        # 3.2 + 1.5
    "method": 2.5,          # 1.0 + 1.5
    "params_count": 3.55,   # 3.4 + 0.15
    "params_values": 5.0,   # 3.5 + 1.5
    "status_codes": 3.25,   # 3.0 + 0.25
    "response": 11.5,       # 10.0 + 1.5
    "confidence": 2.9,      # 2.5 + 0.4
    "confidence_level": 4.2, # 3.5 + 0.7
    "notes": 5.5            # 4.0 + 1.5
}

# Convert cm to Excel units (1 cm = 3.78 Excel units approximately)
COLUMN_WIDTHS = {k: v * 3.78 for k, v in COLUMN_WIDTHS_CM.items()}

# Row heights (in cm, converted to points)
# Excel row height: 1 cm = 28.35 points
HEADER_ROW_HEIGHT_CM = 0.7
DATA_ROW_HEIGHT_CM = 3.85
HEADER_ROW_HEIGHT_POINTS = HEADER_ROW_HEIGHT_CM * 28.35
DATA_ROW_HEIGHT_POINTS = DATA_ROW_HEIGHT_CM * 28.35

# Alignment configuration
CENTER_ALIGNED_COLUMNS = {"method", "params_count", "status_codes", "confidence", "confidence_level"}
TOP_ALIGNED_COLUMNS = {"response"}

_excel_workbook = None
_excel_worksheet = None
_excel_filename = None

def init_excel_file(filename: str):
    """Initialize Excel file with headers and formatting."""
    global _excel_workbook, _excel_worksheet, _excel_filename
    
    if not OPENPYXL_AVAILABLE:
        return
    
    try:
        excel_filename = filename.replace('.csv', '.xlsx')
        _excel_filename = excel_filename
        
        _excel_workbook = Workbook()
        _excel_worksheet = _excel_workbook.active
        _excel_worksheet.title = "Unauth Check Results"
        
        # Write headers with proper alignment
        for col_idx, field in enumerate(CSV_FIELDS, 1):
            cell = _excel_worksheet.cell(row=1, column=col_idx)
            cell.value = field
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Apply alignment based on column type (title row is top-aligned vertically)
            # "response" and "notes" are center-aligned in title row, rest keep their alignment
            if field in {"response", "notes"}:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif field in CENTER_ALIGNED_COLUMNS:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Set column widths (in Excel units, converted from cm)
        for col_idx, field in enumerate(CSV_FIELDS, 1):
            width = COLUMN_WIDTHS.get(field, 15)
            _excel_worksheet.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Set header row height (in points, converted from cm)
        _excel_worksheet.row_dimensions[1].height = HEADER_ROW_HEIGHT_POINTS
        
    except Exception as e:
        print(f"\n[!] Warning: Failed to initialize Excel file: {e}", file=sys.stderr)
        _excel_workbook = None
        _excel_worksheet = None

def append_excel_row(row: Dict[str, Any]):
    """Append a row to the Excel file and save in real-time."""
    global _excel_worksheet, _excel_workbook, _excel_filename
    
    if not OPENPYXL_AVAILABLE or _excel_worksheet is None:
        return
    
    try:
        row_num = _excel_worksheet.max_row + 1
        
        # Set row height for data rows (in points, converted from cm)
        _excel_worksheet.row_dimensions[row_num].height = DATA_ROW_HEIGHT_POINTS
        
        for col_idx, field in enumerate(CSV_FIELDS, 1):
            cell = _excel_worksheet.cell(row=row_num, column=col_idx)
            value = row.get(field, "")
            cell.value = str(value)[:5000] if len(str(value)) > 5000 else value  # Limit to 5000 chars for Excel
            
            # Apply alignment based on column type
            if field in TOP_ALIGNED_COLUMNS:
                # Top-aligned (response column)
                if field in CENTER_ALIGNED_COLUMNS:
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif field in CENTER_ALIGNED_COLUMNS:
                # Center-aligned, vertically centered
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Left-aligned, vertically centered
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Save Excel file in real-time after each row (like CSV)
        if _excel_filename and _excel_workbook:
            _excel_workbook.save(_excel_filename)
        
    except Exception as e:
        print(f"\n[!] Warning: Failed to write Excel row: {e}", file=sys.stderr)

def save_excel_file():
    """Save and close the Excel file."""
    global _excel_workbook, _excel_worksheet, _excel_filename
    
    if not OPENPYXL_AVAILABLE or _excel_workbook is None:
        return
    
    try:
        if _excel_filename:
            _excel_workbook.save(_excel_filename)
            _excel_workbook.close()
    except Exception as e:
        print(f"\n[!] Warning: Failed to save Excel file: {e}", file=sys.stderr)
    finally:
        _excel_workbook = None
        _excel_worksheet = None
        _excel_filename = None

# -------------------------------------------------
# CSV Helpers (Resume-safe)
# -------------------------------------------------

def write_csv_header_if_needed(filename: str):
    if not os.path.exists(filename):
        try:
            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
                writer.writeheader()
            
            # Initialize Excel file if available
            init_excel_file(filename)
        except Exception as e:
            raise ValueError(f"Failed to create CSV file {filename}: {e}")

def append_csv_row(filename: str, row: Dict[str, Any]):
    try:
        with open(filename, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
            writer.writerow(row)
        
        # Also append to Excel file if available
        append_excel_row(row)
    except Exception as e:
        # Log error but don't crash - allow scan to continue
        print(f"\n[!] Warning: Failed to write CSV row: {e}", file=sys.stderr)

def load_completed_endpoints(filename: str) -> set:
    completed = set()
    if not os.path.exists(filename):
        return completed

    try:
        with open(filename, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Include params_values in key to track all 3 test cases
                key = f"{row['method']} {row['endpoint']} {row.get('params_values', '')}"
                completed.add(key)
    except Exception as e:
        # If we can't read the file, assume no endpoints are completed
        # This allows the scan to continue even if CSV is corrupted
        return completed

    return completed

# -------------------------------------------------
# OpenAPI Loading
# -------------------------------------------------

def extract_base_url_from_spec(spec: Dict) -> str:
    """Extract base URL from OpenAPI spec's servers field."""
    # OpenAPI 3.0+ uses 'servers' array
    servers = spec.get("servers", [])
    if servers and isinstance(servers, list) and len(servers) > 0:
        server_url = servers[0].get("url", "") if isinstance(servers[0], dict) else str(servers[0])
        if server_url:
            # Remove trailing slash if present
            return server_url.rstrip("/")
    
    # OpenAPI 2.0 uses 'host', 'basePath', 'schemes'
    host = spec.get("host", "")
    base_path = spec.get("basePath", "")
    schemes = spec.get("schemes", ["https"])
    scheme = schemes[0] if schemes else "https"
    
    if host:
        base_url = f"{scheme}://{host}{base_path}".rstrip("/")
        return base_url
    
    return ""

def load_openapi(url: str = None, file_path: str = None) -> Tuple[Dict, str]:
    spec = None
    
    if url:
        try:
            r = requests.get(url, timeout=15)
            r.raise_for_status()
            spec = r.json()
            # Try to extract from spec first, fallback to URL base
            base_url = extract_base_url_from_spec(spec)
            if not base_url:
                base_url = r.url.rsplit("/", 1)[0]
            return spec, base_url
        except requests.RequestException as e:
            raise ValueError(f"Failed to fetch OpenAPI spec from URL: {e}")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in OpenAPI spec: {e}")

    if file_path:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                spec = json.load(f)
        except FileNotFoundError:
            raise ValueError(f"OpenAPI file not found: {file_path}")
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in OpenAPI file: {e}")
        except Exception as e:
            raise ValueError(f"Error reading OpenAPI file: {e}")
        
        base_url = extract_base_url_from_spec(spec)
        if not base_url:
            raise ValueError(
                "No base URL found in OpenAPI spec. "
                "Please provide 'servers' (OpenAPI 3.0+) or 'host' (OpenAPI 2.0) in the spec, "
                "or use --url instead of --file."
            )
        return spec, base_url

    raise ValueError("No OpenAPI source provided")

# -------------------------------------------------
# Endpoint Extraction
# -------------------------------------------------

def extract_endpoints(spec: Dict) -> List[Dict[str, Any]]:
    endpoints = []

    for path, methods in spec.get("paths", {}).items():
        if not isinstance(methods, dict):
            continue

        for method, details in methods.items():
            if not isinstance(details, dict):
                continue

            params = []
            for p in details.get("parameters", []):
                params.append({
                    "name": p.get("name"),
                    "type": p.get("schema", {}).get("type", "string"),
                    "description": p.get("description", "")
                })

            endpoints.append({
                "path": path,
                "method": method.upper(),
                "params": params
            })

    return endpoints

# -------------------------------------------------
# AI Sample Generation (Max 2 sets)
# -------------------------------------------------

def generate_param_samples(endpoint: Dict[str, Any]) -> List[Dict[str, str]]:
    if not endpoint["params"]:
        return [{}]

    agent = get_agent()
    samples = []
    for _ in range(2):
        sample = {}
        for p in endpoint["params"]:
            sample[p["name"]] = agent.generate_sample_value(
                p["type"], p["name"], p["description"]
            )
        samples.append(sample)

    return samples

# -------------------------------------------------
# Response Body Cleaner
# -------------------------------------------------

def clean_response_body(text: str, limit: int = 500) -> str:
    if not text:
        return "<empty response>"

    text = text.strip()

    try:
        parsed = json.loads(text)
        pretty = json.dumps(parsed, indent=2)
        result = pretty[:limit]
    except Exception:
        result = text[:limit]
    
    # Replace newlines with spaces for CSV compatibility (CSV writer will handle quotes)
    # But keep it readable - we'll let csv.DictWriter handle proper escaping
    return result

# -------------------------------------------------
# Endpoint Testing
# -------------------------------------------------

def format_params_values(params: Dict[str, str]) -> str:
    """Format params dict as JSON string for CSV."""
    if not params:
        return "{}"
    return json.dumps(params, sort_keys=True)

def test_endpoint(endpoint, base_url, csv_file, index, total, verbose):
    # Ensure base_url ends without slash and path starts with slash
    base = base_url.rstrip("/")
    path = endpoint['path'] if endpoint['path'].startswith("/") else f"/{endpoint['path']}"
    url = f"{base}{path}"
    
    params_count = len(endpoint["params"])
    
    if verbose:
        print(f"\n[+] Testing endpoint ({index}/{total}): {endpoint['method']} {endpoint['path']}")
        params_list = [p["name"] for p in endpoint["params"]]
        print(f"    Detected parameters ({params_count}): {', '.join(params_list) if params_list else 'None'}")

    # Generate 2 sets of parameter samples
    samples = generate_param_samples(endpoint)
    
    # Test 3 times: empty params, set 1, set 2
    test_cases = [({}, "empty")]
    
    if samples and len(samples) > 0:
        test_cases.append((samples[0], "set_1"))
    if samples and len(samples) > 1:
        test_cases.append((samples[1], "set_2"))
    
    if verbose:
        print(f"    [*] Agent generating sample values...")
        if params_count > 0:
            print(f"    [*] {len(samples) * params_count} sample values created")
        print(f"    [*] Testing {len(test_cases)} cases: empty params + {len(samples)} sample sets")

    for case_idx, (params, case_name) in enumerate(test_cases, 1):
        statuses = set()
        params_str = format_params_values(params)
        response_body = ""
        
        if verbose:
            if case_name == "empty":
                print(f"    [*] Test case {case_idx}/3: Empty parameters")
            else:
                print(f"    [*] Test case {case_idx}/3: {case_name} - {params_str}")

        try:
            r = requests.request(
                endpoint["method"],
                url,
                params=params,
                timeout=10
            )

            statuses.add(r.status_code)
            # Capture response body (cleaned/truncated for CSV)
            response_body = clean_response_body(r.text, limit=2000)  # Increased limit for CSV

            if verbose:
                print(f"        → Status: {r.status_code}")
                print("        → Response:")
                print(clean_response_body(r.text))

        except requests.RequestException as e:
            if verbose:
                print(f"        [!] Error: {e}")
            statuses.add("ERROR")
            response_body = f"Request Error: {str(e)}"

        confidence = 60 if 200 in statuses else 0
        confidence_level = "Medium" if 200 in statuses else "Inconclusive"

        result = {
            "endpoint": endpoint["path"],
            "method": endpoint["method"],
            "params_count": params_count,
            "params_values": params_str,
            "status_codes": ",".join(map(str, sorted(statuses))) if statuses else "N/A",
            "response": response_body,
            "confidence": confidence,
            "confidence_level": confidence_level,
            "notes": f"Test case: {case_name}"
        }

        append_csv_row(csv_file, result)

    if verbose:
        print("    [+] Endpoint evaluation completed")
        print(f"    [+] {len(test_cases)} result(s) appended to CSV")

# -------------------------------------------------
# File Versioning
# -------------------------------------------------

def extract_hostname_from_url(base_url: str) -> str:
    """Extract hostname from base URL for file naming."""
    try:
        parsed = urlparse(base_url)
        hostname = parsed.hostname or parsed.netloc or "unknown"
        # Remove port if present
        hostname = hostname.split(":")[0]
        # Replace dots and special chars with hyphens
        hostname = re.sub(r'[^a-zA-Z0-9-]', '-', hostname)
        # Remove multiple consecutive hyphens
        hostname = re.sub(r'-+', '-', hostname)
        return hostname.lower()
    except Exception:
        return "unknown-host"

def get_versioned_filename(base_filename: str) -> str:
    """Get next versioned filename (e.g., hostname.csv -> hostname1.csv -> hostname2.csv)."""
    if not os.path.exists(base_filename):
        return base_filename
    
    # Extract base name and extension
    base_name, ext = os.path.splitext(base_filename)
    
    # Find all existing versions in the same directory
    directory = os.path.dirname(base_filename) or "."
    pattern = re.compile(rf"^{re.escape(base_name)}(\d+)?{re.escape(ext)}$")
    
    max_version = -1
    for filename in os.listdir(directory):
        match = pattern.match(filename)
        if match:
            version_str = match.group(1)
            if version_str:
                try:
                    version = int(version_str)
                    max_version = max(max_version, version)
                except ValueError:
                    pass
            else:
                # Base file exists (no number)
                max_version = max(max_version, -1)
    
    # Return next version
    next_version = max_version + 1
    if next_version == 0:
        # Base file exists, return version 1
        return f"{base_name}1{ext}"
    else:
        return f"{base_name}{next_version}{ext}"

# -------------------------------------------------
# Scan Runner (Resume-aware)
# -------------------------------------------------

def run_scan(url=None, file_path=None, output_file=None, verbose=False):
    spec, base_url = load_openapi(url, file_path)
    endpoints = extract_endpoints(spec)
    total = len(endpoints)

    # Generate output filename from hostname if not provided
    if output_file is None:
        hostname = extract_hostname_from_url(base_url)
        output_file = f"{hostname}.csv"
    
    # Get versioned filename
    output_file = get_versioned_filename(output_file)
    
    write_csv_header_if_needed(output_file)
    completed = load_completed_endpoints(output_file)

    if verbose:
        print("[*] OpenAPI loaded successfully")
        print(f"[*] Base URL: {base_url}")
        print(f"[*] Total endpoints detected: {total}")
        print(f"[*] Output file: {output_file}")
        print(f"[*] Resuming scan, {len(completed)} test cases already completed\n")

    completed_count = len(completed)
    total_test_cases = total * 3  # Each endpoint has 3 test cases

    for ep_idx, ep in enumerate(endpoints, 1):
        # Test endpoint with 3 cases (empty, set1, set2)
        # Check if all 3 cases are completed
        empty_key = f"{ep['method']} {ep['path']} {{}}"
        set1_key = None
        set2_key = None
        
        if ep["params"]:
            samples = generate_param_samples(ep)
            if samples:
                set1_key = f"{ep['method']} {ep['path']} {format_params_values(samples[0])}"
            if len(samples) > 1:
                set2_key = f"{ep['method']} {ep['path']} {format_params_values(samples[1])}"
        
        # Check if all test cases are completed
        all_completed = empty_key in completed
        if set1_key:
            all_completed = all_completed and (set1_key in completed)
        if set2_key:
            all_completed = all_completed and (set2_key in completed)
        
        if all_completed:
            # Count completed test cases for progress
            test_case_count = 1 + (1 if set1_key else 0) + (1 if set2_key else 0)
            completed_count += test_case_count
            update_progress(completed_count, total_test_cases)
            continue

        test_endpoint(
            ep,
            base_url,
            output_file,
            ep_idx,
            total,
            verbose
        )
        # Each endpoint adds 3 test cases
        completed_count += 3
        update_progress(completed_count, total_test_cases)

    # Save Excel file if it was created
    excel_file = output_file.replace('.csv', '.xlsx')
    save_excel_file()
    
    if verbose:
        print("\n[*] Scan completed successfully")
        print(f"[*] Results stored in {output_file}")
        if OPENPYXL_AVAILABLE and os.path.exists(excel_file):
            print(f"[*] Excel file with formatting: {excel_file}")
        elif not OPENPYXL_AVAILABLE:
            print("[*] Note: Install 'openpyxl' (pip install openpyxl) for Excel output with proper formatting")
